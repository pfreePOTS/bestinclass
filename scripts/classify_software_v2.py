"""
PulseOne Software COGS & Expense — Three-Bucket Classification (v2)
====================================================================
Uses exact vendor names extracted from the actual workbook.

Three buckets:
  B1 = Client-Delivered Software COGS  (resale to a specific client)
  B2 = Help Desk / Service Delivery COGS (technician delivery tools)
  B3 = Internal SG&A (internal operations)

Source file structure (single sheet "Sheet1"):
  - Col B (index 1): section header OR None
  - Col F (index 5): transaction type (Bill / Deposit) — marks data rows
  - Col H (index 7): date
  - Col L (index 11): vendor name
  - Col V (index 21): amount
  - Col X (index 23): running total formula

Sections:
  Row 2:   "Agreement License Cost"          → rows 3–37
  Row 38:  "Total Agreement License Cost"    → subtotal row
  Row 39:  "Purchases - Software for Resale" → rows 40–474
  Row 475: "Computer and Internet Expenses"  → rows 476–1105
"""

import openpyxl
import json
import os
from collections import defaultdict
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

XLSX_IN  = "/home/ubuntu/bic-work/data/financial/2025 PulseOne Software COGS and Expense.xlsx"
XLSX_OUT = "/home/ubuntu/bic-work/data/financial/Software_COGS_Classified.xlsx"
JSON_OUT = "/home/ubuntu/bic-work/docs/analysis/software_classification_summary.json"

# ── Bucket labels ─────────────────────────────────────────────────────────────
B1 = "Client-Delivered Software COGS"
B2 = "Help Desk / Service Delivery COGS"
B3 = "Internal SG&A"

CONFIRMED = "Confirmed (owner)"
ANALYSIS  = "Analysis"
INFERRED  = "Inferred — needs confirmation"

# ── Master classification table ───────────────────────────────────────────────
# Keyed by EXACT vendor name from the workbook (case-insensitive match)
# Value: (bucket, confidence, rationale)
CLASSIFICATIONS = {
    # ════════════════════════════════════════════════════════════════════════
    # SECTION 1: Agreement License Cost  ($373,999 total)
    # ════════════════════════════════════════════════════════════════════════

    # Synnex Corp — $301,813 in Agreement License Cost
    # Consolidated "Estimate" billing with no line-item detail.
    # Includes managed-service software bundles (Barracuda, backup, security)
    # billed to PulseOne as a distributor. Classified B2 because these are
    # tools used to deliver managed services, not standalone resale.
    # ⚠ HIGH PRIORITY: No client attribution. Monthly variance 44.5%.
    "synnex corp": (B2, INFERRED,
        "Agreement License Cost — consolidated distributor bill for managed-service software stack. "
        "No line-item detail available. Monthly variance 44.5% ($20,947–$30,277) unexplained. "
        "Needs ConnectWise agreement cross-reference to confirm client attribution."),

    # Barracuda (Skout Cybersecurity) — $72,960
    # Managed email security / EDR platform delivered to clients.
    # 87% cost escalation Jan→Dec 2025 ($3,996→$7,472/mo).
    "barracuda (skout cybersecurity)": (B2, ANALYSIS,
        "Managed security delivery platform (email filtering + EDR). "
        "Cost grew 87% in CY2025 — $41,713/year run-rate increase. "
        "Confirm all cost increases are passed through to client billing."),

    # Amazon Web Services — $10,709 in Agreement License Cost
    # Owner confirmed: client-specific AWS environments bundled in agreements.
    "amazon web services": (B2, CONFIRMED,
        "Client-specific AWS environments bundled in managed service agreements. "
        "Owner confirmed COGS April 20, 2026."),

    # Microsoft — negative $11,482 (rebates/credits)
    # Rebates from Microsoft partner program applied against Agreement License Cost.
    "microsoft": (B2, ANALYSIS,
        "Microsoft partner rebates/credits applied against Agreement License Cost. "
        "Negative amount reduces COGS. Correct classification."),

    # ════════════════════════════════════════════════════════════════════════
    # SECTION 2: Purchases - Software for Resale  ($570,987 total)
    # ════════════════════════════════════════════════════════════════════════

    # Synnex Corp — $251,028 in Software Resale
    # Standalone product purchases billed to specific clients.
    # This is the primary driver of the -9.4% software resale margin.
    "synnex corp_resale": (B1, ANALYSIS,
        "Standalone software/hardware resale via TD Synnex distributor. "
        "Primary driver of -9.4% software resale margin. "
        "Client-specific purchases — must be repriced to 25%+ margin."),

    # Ingram Micro — $194,247
    # Standalone product purchases. ARC alone = $114,496 (57% of Ingram spend).
    "ingram micro": (B1, ANALYSIS,
        "Standalone software/hardware resale via Ingram Micro distributor. "
        "ARC client = $114,496 (57% of total Ingram spend). "
        "Repricing ARC alone recovers ~$48K — eliminates entire segment loss."),

    # Amazon Web Services — $63,273 in Software Resale
    # Owner confirmed: client-specific AWS. Separate from the $10,709 in Agreement Licenses.
    "amazon web services_resale": (B1, CONFIRMED,
        "Client-specific AWS consumption billed as standalone resale. "
        "Owner confirmed COGS April 20, 2026. Separate from agreement-bundled AWS."),

    # AvePoint — $41,200
    # Microsoft 365 data governance/backup platform. Resold to clients.
    "avepoint": (B1, ANALYSIS,
        "Microsoft 365 backup and governance platform resold to clients. "
        "High-value resale item — verify client billing covers 25%+ margin."),

    # EFolder Systems (Axcient) — $10,719
    # Business continuity / backup platform resold to clients.
    "efolder systems (axcient)": (B1, ANALYSIS,
        "Business continuity and backup platform (Axcient) resold to clients."),

    # ScalePad {WarrantyMaster} — $5,988
    # Hardware warranty management resold to clients.
    "scalepad {warrantymaster}": (B1, ANALYSIS,
        "Hardware lifecycle and warranty management platform resold to clients."),

    # Atlassian — $2,962 in Software Resale
    # Jira/Confluence. Could be internal (SGA) or client-facing (COGS).
    # In the Resale section → likely resold to a specific client.
    "atlassian": (B1, INFERRED,
        "Jira/Confluence in Resale section — likely resold to a specific client. "
        "Confirm whether this is a client resale or internal tool miscategorized."),

    # Mosyle Corp — $600
    # Apple MDM platform. Likely resold to Mac-heavy client.
    "mosyle corp": (B1, ANALYSIS,
        "Apple MDM platform (Mosyle) — likely resold to Mac-heavy client."),

    # AppRiver — $453
    # Email security/Microsoft 365 distribution. Resale.
    "appriver": (B1, ANALYSIS,
        "Email security and Microsoft 365 distribution resale via AppRiver."),

    # Paddle.net — $449
    # Software licensing platform. Likely a specific software title resold.
    "paddle.net": (B1, INFERRED,
        "Software licensing via Paddle.net — likely a specific title resold to client. "
        "Confirm which product and client."),

    # Microsoft (in Resale section) — $69
    # Small Microsoft licensing item in resale section.
    "microsoft_resale": (B1, ANALYSIS,
        "Small Microsoft licensing item in Software Resale section."),

    # ════════════════════════════════════════════════════════════════════════
    # SECTION 3: Computer and Internet Expenses  ($256,232 total)
    # ════════════════════════════════════════════════════════════════════════

    # ConnectWise, Inc. — $89,800
    # PSA + RMM platform. Core service delivery infrastructure.
    # Industry standard: 65% COGS / 35% SGA split for PSA/RMM.
    # Currently in SGA — should be reclassified to B2 (COGS).
    "connectwise, inc.": (B2, ANALYSIS,
        "PSA/RMM platform — primary service delivery infrastructure. "
        "Currently in SGA; should be reclassified to Service Delivery COGS. "
        "Industry standard: 65% COGS / 35% SGA. Full $89,800 moved to COGS "
        "given PulseOne's delivery-centric use."),

    # Flexis — $64,009
    # NOC/Help Desk augmentation. Owner confirmed COGS April 20, 2026.
    "flexis": (B2, CONFIRMED,
        "NOC/Help Desk augmentation service. Owner confirmed COGS April 20, 2026. "
        "Currently in SGA — must be reclassified to Service Delivery COGS."),

    # Microsoft Azure — $19,041
    # Internal Azure infrastructure. Owner indicated AWS is client-specific COGS;
    # Azure here is likely internal PulseOne infrastructure (SGA).
    # ⚠ Needs confirmation — could be client-hosted environments.
    "microsoft azure": (B3, INFERRED,
        "Microsoft Azure in Computer & Internet SGA. "
        "If PulseOne internal infrastructure → SGA correct. "
        "If client-specific hosted environments → reclassify to B2 COGS. "
        "Needs owner confirmation."),

    # Microsoft — $14,127 in SGA
    # Internal Microsoft 365 / productivity licenses for PulseOne staff.
    "microsoft_sga": (B3, ANALYSIS,
        "Internal Microsoft 365 / productivity licenses for PulseOne staff. SGA correct."),

    # Kaseya — $10,563
    # RMM/security platform. In SGA — should be Service Delivery COGS.
    "kaseya": (B2, ANALYSIS,
        "RMM and security management platform used for client service delivery. "
        "Currently in SGA — should be reclassified to Service Delivery COGS."),

    # Samurai Sync — $9,450
    # Owner confirmed COGS April 20, 2026. Service delivery tool.
    "samurai sync": (B2, CONFIRMED,
        "Service delivery tool. Owner confirmed COGS April 20, 2026. "
        "Currently in SGA — must be reclassified to Service Delivery COGS."),

    # Wasabi Technologies — $7,353
    # Cloud backup storage for client data. Service delivery.
    "wasabi technologies": (B2, ANALYSIS,
        "Cloud backup storage for client data delivery. "
        "Currently in SGA — should be reclassified to Service Delivery COGS."),

    # Avalara — $6,807
    # Tax compliance automation. Internal finance/accounting tool.
    "avalara": (B3, ANALYSIS,
        "Tax compliance and sales tax automation. Internal finance tool. SGA correct."),

    # MSP360 — $4,005
    # Backup management platform. Used for client backup delivery.
    "msp360": (B2, ANALYSIS,
        "Backup management platform for client data protection delivery. "
        "Currently in SGA — should be reclassified to Service Delivery COGS."),

    # Adobe Software — $3,739
    # Creative suite. Internal marketing/design tool.
    "adobe software": (B3, ANALYSIS,
        "Adobe Creative Suite — internal marketing and design tool. SGA correct."),

    # TimeZest — $3,609
    # ConnectWise scheduling integration. Borderline COGS/SGA.
    # Used by technicians to schedule client appointments → lean COGS.
    "timezest": (B2, ANALYSIS,
        "ConnectWise scheduling integration used by technicians for client appointments. "
        "Borderline COGS/SGA — classified as Service Delivery COGS given direct client interaction."),

    # CloudBase Services — $3,600
    # Cloud services. Needs clarification — could be client or internal.
    "cloudbase services": (B2, INFERRED,
        "Cloud services — likely client delivery infrastructure. "
        "Confirm whether client-specific or internal PulseOne infrastructure."),

    # GoDaddy.com, Inc. — $2,880
    # Domain registration and web hosting. Likely internal (PulseOne website/email).
    # Could include client-hosted domains — needs confirmation.
    "godaddy.com, inc.": (B3, INFERRED,
        "Domain registration and web hosting. Likely internal PulseOne domains. "
        "Confirm whether any domains are client-hosted (would be B1 resale)."),

    # CyberFox — $2,189
    # Cybersecurity tool. Likely client security delivery.
    "cyberfox": (B2, ANALYSIS,
        "Cybersecurity platform — likely used for client security delivery."),

    # Duo Security — $2,068
    # MFA platform for client access security. Service delivery.
    "duo security": (B2, ANALYSIS,
        "Client MFA/access security delivery platform. "
        "Currently in SGA — should be reclassified to Service Delivery COGS."),

    # Pivotal Crew LLC — $2,000
    # One-time charge. Likely a contractor or specialized service.
    "pivotal crew llc": (B3, INFERRED,
        "One-time charge — likely a contractor or specialized service. "
        "Confirm purpose: if client delivery work → B2 COGS; if internal → B3 SGA."),

    # Cursor — $1,184
    # AI coding IDE. Internal development tool.
    "cursor": (B3, ANALYSIS,
        "AI coding IDE — internal development tool. SGA correct."),

    # OpenAI — $1,076
    # AI API. Internal tool (AI assistant, automation).
    "openai": (B3, ANALYSIS,
        "OpenAI API — internal AI tooling. SGA correct. "
        "Review for consolidation with other AI subscriptions."),

    # Qualys — $1,024
    # Vulnerability scanning platform for client security delivery.
    "qualys": (B2, ANALYSIS,
        "Client vulnerability scanning platform. "
        "Currently in SGA — should be reclassified to Service Delivery COGS."),

    # Eversign — $960
    # E-signature platform. Internal contracts/agreements.
    "eversign": (B3, ANALYSIS,
        "E-signature platform for internal contracts and client agreements. SGA correct."),

    # Railway — $823
    # Internal hosting/deployment platform for dev projects.
    "railway": (B3, ANALYSIS,
        "Internal hosting/deployment platform. SGA correct."),

    # Replit, Inc. — $526
    # Online coding environment. Internal dev tool.
    "replit, inc.": (B3, ANALYSIS,
        "Online coding environment — internal development tool. SGA correct. "
        "Redundant with Cursor — candidate for cancellation."),

    # Calendy — $514 (note: spelled 'Calendy' in the data, not 'Calendly')
    "calendy": (B3, ANALYSIS,
        "Scheduling tool — internal use. SGA correct. "
        "Redundant with TimeZest — candidate for cancellation."),

    # Anthropic — $425
    # Claude AI API. Internal AI tool.
    "anthropic": (B3, ANALYSIS,
        "Anthropic Claude API — internal AI tooling. SGA correct. "
        "Redundant with OpenAI — consolidate to one primary LLM."),

    # Manus AI — $389
    # AI assistant platform (this audit tool). Internal.
    "manus ai": (B3, ANALYSIS,
        "AI assistant platform — internal strategic analysis tool. SGA correct."),

    # AI MAD FZCO — $360
    # AI tool/service. Internal.
    "ai mad fzco": (B3, ANALYSIS,
        "AI service — internal tool. SGA correct. Confirm purpose and necessity."),

    # Zapier.com — $360
    # Workflow automation. Internal operations.
    "zapier.com": (B3, ANALYSIS,
        "Workflow automation — internal operations tool. SGA correct."),

    # Google — $352
    # Google Workspace or Google Cloud. Internal productivity.
    "google": (B3, ANALYSIS,
        "Google services (Workspace/Cloud) — internal productivity. SGA correct."),

    # Monday.com — $288
    # Project management. Internal.
    "monday.com": (B3, ANALYSIS,
        "Project management platform — internal operations. SGA correct."),

    # RapidAPI — $261
    # API marketplace. Internal dev/integration tool.
    "rapidapi": (B3, ANALYSIS,
        "API marketplace — internal development tool. SGA correct."),

    # Iron Software — $249
    # .NET library for document processing. Internal dev.
    "iron software": (B3, ANALYSIS,
        "Document processing library — internal development tool. SGA correct."),

    # Scrapfly — $230
    # Web scraping service. Internal data collection.
    "scrapfly": (B3, ANALYSIS,
        "Web scraping service — internal data collection tool. SGA correct."),

    # Claude.AI — $220
    # Claude AI web interface. Internal. Redundant with Anthropic API.
    "claude.ai": (B3, ANALYSIS,
        "Claude AI web interface — internal tool. SGA correct. "
        "Redundant with Anthropic API subscription — consolidate."),

    # Frame.io — $210
    # Video review platform. Internal creative/marketing.
    "frame.io": (B3, ANALYSIS,
        "Video review and collaboration platform — internal marketing tool. SGA correct."),

    # Cognito LLC — $204
    # Forms/data collection. Internal or client-facing surveys.
    "cognito llc": (B3, INFERRED,
        "Forms and data collection platform. "
        "Confirm if used for client-facing forms (borderline COGS) or internal only."),

    # LearnDash — $199
    # LMS platform. Internal training.
    "learndash": (B3, ANALYSIS,
        "Learning Management System — internal staff training. SGA correct."),

    # Vevox — $131
    # Meeting polling/Q&A tool. Internal meetings.
    "vevox": (B3, ANALYSIS,
        "Meeting polling tool — internal use. SGA correct. Candidate for cancellation."),

    # Zoho — $120
    # CRM/productivity suite. Internal.
    "zoho": (B3, ANALYSIS,
        "Zoho CRM/productivity suite — internal tool. SGA correct. "
        "Confirm if actively used or can be cancelled."),

    # PromptSMRT — $109
    # AI prompt tool. Internal.
    "promptsmrt": (B3, ANALYSIS,
        "AI prompt management tool — internal. SGA correct."),

    # Reincubate — $100
    # iPhone backup recovery tool. Internal IT.
    "reincubate": (B3, ANALYSIS,
        "iPhone backup recovery tool — internal IT use. SGA correct."),

    # Figma — $97
    # Design tool. Internal marketing/design.
    "figma": (B3, ANALYSIS,
        "Design tool — internal marketing use. SGA correct."),

    # Carbonite — $96
    # Backup software. Could be client delivery or internal.
    "carbonite": (B3, INFERRED,
        "Backup software — confirm if for client delivery (B2 COGS) or internal backup (B3 SGA)."),

    # Aqua Voice — $90
    # Voice/transcription tool. Internal.
    "aqua voice": (B3, ANALYSIS,
        "Voice transcription tool — internal use. SGA correct."),

    # Quicken — $72
    # Personal finance software. Non-business tool.
    "quicken": (B3, ANALYSIS,
        "Personal finance software — non-business use. SGA. Candidate for immediate cancellation."),

    # Patreon — $71
    # Content subscription platform. Non-business.
    "patreon": (B3, ANALYSIS,
        "Content subscription platform — non-business use. Candidate for immediate cancellation."),

    # Wordpress Rocket — $59
    # WordPress caching plugin. Internal website.
    "wordpress rocket": (B3, ANALYSIS,
        "WordPress caching plugin — internal website. SGA correct."),

    # SessionBox — $48
    # Multi-session browser tool. Internal dev/testing.
    "sessionbox": (B3, ANALYSIS,
        "Multi-session browser tool — internal development/testing. SGA correct."),

    # DiviBooster — $39
    # WordPress plugin. Internal website.
    "divibooster": (B3, ANALYSIS,
        "WordPress plugin — internal website. SGA correct."),

    # OpenRouter — $27
    # AI model routing. Internal. Redundant with OpenAI/Anthropic.
    "openrouter": (B3, ANALYSIS,
        "AI model routing service — internal. Redundant with direct API subscriptions. "
        "Candidate for cancellation."),

    # Google Storage — $24
    # Google Cloud Storage. Internal.
    "google storage": (B3, ANALYSIS,
        "Google Cloud Storage — internal data storage. SGA correct."),

    # Have I Been Pwned — $22
    # Breach monitoring API. Could be client security delivery.
    "have i been pwned": (B2, ANALYSIS,
        "Breach monitoring API — likely used for client security monitoring delivery."),

    # Cloudfare — $21 (note: spelled 'Cloudfare' in data, not 'Cloudflare')
    "cloudfare": (B3, ANALYSIS,
        "Cloudflare CDN/security — internal website infrastructure. SGA correct."),

    # Event Connect — $14
    # Event management tool. Internal.
    "event connect": (B3, ANALYSIS,
        "Event management tool — internal use. SGA correct."),
}

# ── Section-to-default-bucket mapping ────────────────────────────────────────
SECTION_DEFAULTS = {
    "Agreement License Cost":          (B2, INFERRED, "Agreement License Cost section default — service delivery tool"),
    "Total Agreement License Cost":    (B2, INFERRED, "Subtotal row"),
    "Purchases - Software for Resale": (B1, INFERRED, "Software Resale section default — client-delivered product"),
    "Computer and Internet Expenses":  (B3, INFERRED, "Computer & Internet SGA section default — internal tool"),
}

# Vendors that appear in multiple sections need section-aware lookup
MULTI_SECTION_VENDORS = {
    "amazon web services": {
        "Agreement License Cost": CLASSIFICATIONS["amazon web services"],
        "Purchases - Software for Resale": CLASSIFICATIONS["amazon web services_resale"],
    },
    "synnex corp": {
        "Agreement License Cost": CLASSIFICATIONS["synnex corp"],
        "Purchases - Software for Resale": CLASSIFICATIONS.get("synnex corp_resale",
            (B1, ANALYSIS, "Synnex Corp resale — client-delivered product")),
    },
    "microsoft": {
        "Agreement License Cost": CLASSIFICATIONS["microsoft"],
        "Purchases - Software for Resale": CLASSIFICATIONS.get("microsoft_resale",
            (B1, ANALYSIS, "Microsoft licensing resale")),
        "Computer and Internet Expenses": CLASSIFICATIONS.get("microsoft_sga",
            (B3, ANALYSIS, "Internal Microsoft 365 licenses")),
    },
    "atlassian": {
        "Purchases - Software for Resale": CLASSIFICATIONS["atlassian"],
        "Computer and Internet Expenses": (B3, INFERRED,
            "Atlassian in SGA — likely internal Jira/Confluence. Confirm."),
    },
}


def get_classification(vendor: str, section: str) -> tuple:
    """Return (bucket, confidence, rationale) for a vendor+section combination."""
    if not vendor:
        return SECTION_DEFAULTS.get(section, (B3, INFERRED, "Unknown"))

    vl = vendor.lower().strip()

    # Check multi-section vendors first
    for key, section_map in MULTI_SECTION_VENDORS.items():
        if key in vl:
            if section in section_map:
                return section_map[section]
            # Fall through to general lookup

    # Check general vendor rules (longest match wins)
    best_len = 0
    best_rule = None
    for key, rule in CLASSIFICATIONS.items():
        # Skip the _resale and _sga suffixed keys in general lookup
        if "_resale" in key or "_sga" in key:
            continue
        if key in vl and len(key) > best_len:
            best_len = len(key)
            best_rule = rule

    if best_rule:
        return best_rule

    # Fall back to section default
    return SECTION_DEFAULTS.get(section, (B3, INFERRED, "No match — defaulted"))


def run():
    wb_in = openpyxl.load_workbook(XLSX_IN)
    ws_in = wb_in["Sheet1"]

    # ── Build classified data ─────────────────────────────────────────────────
    rows_out = []
    section = None
    totals = {B1: 0.0, B2: 0.0, B3: 0.0}
    counts = {B1: 0, B2: 0, B3: 0}
    vendor_totals = {B1: defaultdict(float), B2: defaultdict(float), B3: defaultdict(float)}
    section_totals = defaultdict(lambda: {B1: 0.0, B2: 0.0, B3: 0.0, "total": 0.0})
    needs_confirmation = []

    for i, row in enumerate(ws_in.iter_rows(min_row=1, max_row=ws_in.max_row, values_only=True), 1):
        b = row[1]   # section header
        f = row[5]   # transaction type
        dt = row[7]  # date
        vendor = row[11]  # vendor name
        memo = row[13]    # memo/description
        amount = row[21]  # amount

        # Section header row
        if b and not f:
            section = str(b)
            continue

        # Skip non-data rows
        if not f or not vendor:
            continue

        vendor_str = str(vendor).strip()
        amount_val = float(amount) if isinstance(amount, (int, float)) else 0.0

        bucket, confidence, rationale = get_classification(vendor_str, section or "")

        rows_out.append({
            "row": i,
            "section": section,
            "type": f,
            "date": dt,
            "vendor": vendor_str,
            "memo": str(memo).strip() if memo else "",
            "amount": amount_val,
            "bucket": bucket,
            "confidence": confidence,
            "rationale": rationale,
        })

        totals[bucket] += amount_val
        counts[bucket] += 1
        vendor_totals[bucket][vendor_str] += amount_val
        section_totals[section][bucket] += amount_val
        section_totals[section]["total"] += amount_val

        if confidence == INFERRED:
            needs_confirmation.append({
                "vendor": vendor_str,
                "section": section,
                "amount": amount_val,
                "bucket": bucket,
                "rationale": rationale,
            })

    grand_total = sum(totals.values())

    # ── Build output workbook ─────────────────────────────────────────────────
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)

    # Styles
    hdr_fill  = PatternFill("solid", fgColor="1F4E79")
    hdr_font  = Font(color="FFFFFF", bold=True, size=10)
    b1_fill   = PatternFill("solid", fgColor="E2EFDA")   # green — resale
    b2_fill   = PatternFill("solid", fgColor="DDEBF7")   # blue — delivery
    b3_fill   = PatternFill("solid", fgColor="FFF2CC")   # yellow — SGA
    inf_fill  = PatternFill("solid", fgColor="FCE4D6")   # orange — needs confirmation
    sum_fill  = PatternFill("solid", fgColor="D6E4F0")   # light blue — summary
    bold_font = Font(bold=True, size=10)
    center    = Alignment(horizontal="center")
    right     = Alignment(horizontal="right")

    bucket_fills = {B1: b1_fill, B2: b2_fill, B3: b3_fill}

    # ── Sheet 1: SUMMARY ─────────────────────────────────────────────────────
    ws_sum = wb_out.create_sheet("SUMMARY")

    ws_sum.column_dimensions["A"].width = 45
    ws_sum.column_dimensions["B"].width = 14
    ws_sum.column_dimensions["C"].width = 14
    ws_sum.column_dimensions["D"].width = 14

    def sum_row(vals, fill=None, bold=False):
        r = ws_sum.max_row + 1
        ws_sum.append(vals)
        row_ref = ws_sum[r]
        if fill:
            for c in row_ref:
                c.fill = fill
        if bold:
            for c in row_ref:
                c.font = Font(bold=True, size=10)

    ws_sum.append(["PulseOne — Software COGS & Expense: Three-Bucket Classification"])
    ws_sum[1][0].font = Font(bold=True, size=13)
    ws_sum.append(["Source: 2025 PulseOne Software COGS and Expense.xlsx | Classified: April 20, 2026"])
    ws_sum.append([])

    ws_sum.append(["BUCKET", "TRANSACTIONS", "TOTAL AMOUNT", "% OF TOTAL"])
    for c in ws_sum[ws_sum.max_row]:
        c.fill = hdr_fill; c.font = hdr_font; c.alignment = center

    for b in [B1, B2, B3]:
        ws_sum.append([b, counts[b], round(totals[b], 2),
                       f"{totals[b]/grand_total*100:.1f}%" if grand_total else "0%"])
        for c in ws_sum[ws_sum.max_row]:
            c.fill = bucket_fills[b]

    ws_sum.append(["GRAND TOTAL", sum(counts.values()), round(grand_total, 2), "100.0%"])
    for c in ws_sum[ws_sum.max_row]:
        c.font = bold_font

    ws_sum.append([])
    ws_sum.append(["ITEMS NEEDING OWNER CONFIRMATION", len(needs_confirmation), "", ""])
    ws_sum.append([])

    # Section breakdown
    ws_sum.append(["BY SECTION", "", "", ""])
    for c in ws_sum[ws_sum.max_row]:
        c.fill = hdr_fill; c.font = hdr_font

    for sec, sdata in section_totals.items():
        ws_sum.append([sec, "", round(sdata["total"], 2), ""])
        for b in [B1, B2, B3]:
            if sdata[b] != 0:
                ws_sum.append([f"  → {b}", "", round(sdata[b], 2),
                               f"{sdata[b]/sdata['total']*100:.1f}%" if sdata["total"] else ""])
                for c in ws_sum[ws_sum.max_row]:
                    c.fill = bucket_fills[b]

    ws_sum.append([])

    # Top vendors per bucket
    ws_sum.append(["TOP VENDORS BY BUCKET", "", "", ""])
    for c in ws_sum[ws_sum.max_row]:
        c.fill = hdr_fill; c.font = hdr_font

    for b in [B1, B2, B3]:
        ws_sum.append([f"── {b} ──", "", "", ""])
        for c in ws_sum[ws_sum.max_row]:
            c.fill = bucket_fills[b]; c.font = bold_font
        sorted_v = sorted(vendor_totals[b].items(), key=lambda x: x[1], reverse=True)[:15]
        for vname, vamt in sorted_v:
            ws_sum.append(["", vname, round(vamt, 2), f"{vamt/grand_total*100:.1f}%"])
        ws_sum.append([])

    # ── Sheet 2: ALL TRANSACTIONS ─────────────────────────────────────────────
    ws_all = wb_out.create_sheet("ALL TRANSACTIONS")
    ws_all.column_dimensions["A"].width = 35
    ws_all.column_dimensions["B"].width = 10
    ws_all.column_dimensions["C"].width = 12
    ws_all.column_dimensions["D"].width = 35
    ws_all.column_dimensions["E"].width = 30
    ws_all.column_dimensions["F"].width = 14
    ws_all.column_dimensions["G"].width = 42
    ws_all.column_dimensions["H"].width = 22
    ws_all.column_dimensions["I"].width = 70

    headers = ["SECTION", "TYPE", "DATE", "VENDOR", "MEMO", "AMOUNT", "BUCKET", "CONFIDENCE", "RATIONALE"]
    ws_all.append(headers)
    for c in ws_all[1]:
        c.fill = hdr_fill; c.font = hdr_font; c.alignment = center

    for r in rows_out:
        ws_all.append([
            r["section"], r["type"],
            r["date"].strftime("%Y-%m-%d") if hasattr(r["date"], "strftime") else str(r["date"]),
            r["vendor"], r["memo"], round(r["amount"], 2),
            r["bucket"], r["confidence"], r["rationale"]
        ])
        fill = inf_fill if r["confidence"] == INFERRED else bucket_fills.get(r["bucket"], PatternFill())
        for c in ws_all[ws_all.max_row]:
            c.fill = fill

    # ── Sheet 3: NEEDS CONFIRMATION ──────────────────────────────────────────
    ws_conf = wb_out.create_sheet("NEEDS CONFIRMATION")
    ws_conf.column_dimensions["A"].width = 35
    ws_conf.column_dimensions["B"].width = 35
    ws_conf.column_dimensions["C"].width = 14
    ws_conf.column_dimensions["D"].width = 42
    ws_conf.column_dimensions["E"].width = 70

    ws_conf.append(["SECTION", "VENDOR", "AMOUNT", "CURRENT BUCKET", "RATIONALE / QUESTION"])
    for c in ws_conf[1]:
        c.fill = hdr_fill; c.font = hdr_font; c.alignment = center

    seen = set()
    for item in needs_confirmation:
        key = (item["vendor"], item["section"])
        if key not in seen:
            seen.add(key)
            ws_conf.append([item["section"], item["vendor"],
                            round(item["amount"], 2), item["bucket"], item["rationale"]])
            for c in ws_conf[ws_conf.max_row]:
                c.fill = inf_fill

    # ── Sheet 4: VENDOR ROLLUP ────────────────────────────────────────────────
    ws_vend = wb_out.create_sheet("VENDOR ROLLUP")
    ws_vend.column_dimensions["A"].width = 35
    ws_vend.column_dimensions["B"].width = 42
    ws_vend.column_dimensions["C"].width = 14
    ws_vend.column_dimensions["D"].width = 22
    ws_vend.column_dimensions["E"].width = 14

    ws_vend.append(["SECTION", "VENDOR", "TOTAL AMOUNT", "BUCKET", "CONFIDENCE"])
    for c in ws_vend[1]:
        c.fill = hdr_fill; c.font = hdr_font; c.alignment = center

    # Build vendor+section rollup
    vend_rollup = defaultdict(lambda: {"amount": 0.0, "bucket": "", "confidence": "", "section": ""})
    for r in rows_out:
        key = (r["section"], r["vendor"])
        vend_rollup[key]["amount"] += r["amount"]
        vend_rollup[key]["bucket"] = r["bucket"]
        vend_rollup[key]["confidence"] = r["confidence"]
        vend_rollup[key]["section"] = r["section"]

    sorted_rollup = sorted(vend_rollup.items(), key=lambda x: x[1]["amount"], reverse=True)
    for (sec, vname), vdata in sorted_rollup:
        ws_vend.append([sec, vname, round(vdata["amount"], 2), vdata["bucket"], vdata["confidence"]])
        fill = inf_fill if vdata["confidence"] == INFERRED else bucket_fills.get(vdata["bucket"], PatternFill())
        for c in ws_vend[ws_vend.max_row]:
            c.fill = fill

    # Save workbook
    os.makedirs(os.path.dirname(XLSX_OUT), exist_ok=True)
    wb_out.save(XLSX_OUT)
    print(f"✓ Saved: {XLSX_OUT}")

    # ── Save JSON summary ─────────────────────────────────────────────────────
    summary = {
        "classified_date": "2026-04-20",
        "source_file": "2025 PulseOne Software COGS and Expense.xlsx",
        "grand_total": round(grand_total, 2),
        "total_transactions": len(rows_out),
        "by_bucket": {
            b: {
                "count": counts[b],
                "amount": round(totals[b], 2),
                "pct_of_total": round(totals[b]/grand_total*100, 1) if grand_total else 0,
                "top_vendors": [
                    {"vendor": v, "amount": round(a, 2)}
                    for v, a in sorted(vendor_totals[b].items(), key=lambda x: x[1], reverse=True)[:20]
                ]
            }
            for b in [B1, B2, B3]
        },
        "by_section": {
            sec: {
                "total": round(sdata["total"], 2),
                B1: round(sdata[B1], 2),
                B2: round(sdata[B2], 2),
                B3: round(sdata[B3], 2),
            }
            for sec, sdata in section_totals.items()
        },
        "needs_confirmation": [
            {"vendor": item["vendor"], "section": item["section"],
             "amount": round(item["amount"], 2), "bucket": item["bucket"]}
            for item in needs_confirmation
        ]
    }

    os.makedirs(os.path.dirname(JSON_OUT), exist_ok=True)
    with open(JSON_OUT, "w") as f:
        json.dump(summary, f, indent=2)
    print(f"✓ Saved: {JSON_OUT}")

    # ── Console output ────────────────────────────────────────────────────────
    print("\n" + "="*72)
    print("THREE-BUCKET CLASSIFICATION SUMMARY")
    print("="*72)
    print(f"{'Source:':<30} 2025 PulseOne Software COGS and Expense.xlsx")
    print(f"{'Total transactions:':<30} {len(rows_out)}")
    print(f"{'Grand total:':<30} ${grand_total:,.2f}")
    print()

    for b in [B1, B2, B3]:
        pct = totals[b]/grand_total*100 if grand_total else 0
        print(f"\n{'─'*72}")
        print(f"  {b}")
        print(f"  Transactions: {counts[b]}  |  Amount: ${totals[b]:,.2f}  |  {pct:.1f}% of total")
        sorted_v = sorted(vendor_totals[b].items(), key=lambda x: x[1], reverse=True)[:10]
        for vname, vamt in sorted_v:
            print(f"    {vname:<48} ${vamt:>12,.2f}")

    print(f"\n{'─'*72}")
    print(f"\nItems needing owner confirmation: {len(set((x['vendor'],x['section']) for x in needs_confirmation))}")
    print("\nTop items needing confirmation:")
    seen2 = set()
    for item in sorted(needs_confirmation, key=lambda x: abs(x["amount"]), reverse=True):
        key = (item["vendor"], item["section"])
        if key not in seen2:
            seen2.add(key)
            print(f"  [{item['section'][:25]:<25}] {item['vendor']:<40} ${item['amount']:>10,.2f}")
        if len(seen2) >= 10:
            break

    return summary


if __name__ == "__main__":
    result = run()
