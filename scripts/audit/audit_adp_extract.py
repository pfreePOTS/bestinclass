#!/usr/bin/env python3
"""
Extract all employee data from ADP file.
Header row = 7
Data rows = 8 onwards (until row 36 based on earlier scan)
Key columns:
  1: Employee Name
  5: Department
  66: Total Hours
  67: Total Earnings (GROSS)
  85: Total Taxes (employee side)
  91: Net Pay
  104: Total Employer Liability
  105+: Payment details (type, check date, transaction, amount) x 42 payments
"""
import openpyxl
from datetime import datetime

wb = openpyxl.load_workbook(
    "/home/ubuntu/projects/pulseone-best-in-class-audit-da89126e/2025payrollreports.xlsx",
    data_only=True
)
ws = wb[wb.sheetnames[0]]

print("=" * 130)
print("  ADP PAYROLL DETAIL EXTRACT — CY 2025")
print(f"  Check Dates: {ws.cell(3,1).value}")
print(f"  To: {ws.cell(4,1).value}")
print("=" * 130)

def safe_float(val):
    if val is None or val == '' or val == '\xa0':
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0

# Extract per-employee data
employees = []
for row_idx in range(8, ws.max_row + 1):
    name = ws.cell(row_idx, 1).value
    if not name or 'Total' in str(name) or 'Frequency' in str(name):
        continue
    
    dept = ws.cell(row_idx, 5).value or ''
    total_hours = safe_float(ws.cell(row_idx, 66).value)
    total_earnings = safe_float(ws.cell(row_idx, 67).value)
    total_taxes = safe_float(ws.cell(row_idx, 85).value)
    net_pay = safe_float(ws.cell(row_idx, 91).value)
    employer_liability = safe_float(ws.cell(row_idx, 104).value)
    
    # Extract all payment check dates to verify date range
    payment_dates = []
    payment_amounts = []
    for p in range(1, 43):  # Payments 1-42
        base_col = 105 + (p - 1) * 4  # Each payment has 4 columns
        if base_col + 1 <= ws.max_column:
            check_date = ws.cell(row_idx, base_col + 1).value  # Check Date column
            pay_amt = safe_float(ws.cell(row_idx, base_col + 3).value)  # Amount column
            if check_date:
                payment_dates.append(str(check_date))
                payment_amounts.append(pay_amt)
    
    employees.append({
        'name': str(name).strip(),
        'dept': str(dept).strip(),
        'hours': total_hours,
        'gross': total_earnings,
        'ee_taxes': total_taxes,
        'net': net_pay,
        'er_liability': employer_liability,
        'payment_dates': payment_dates,
        'payment_amounts': payment_amounts,
    })

# Print summary
print(f"\n  {'Employee Name':<35} {'Dept':<30} {'Hours':>7} {'Gross':>12} {'EE Tax':>10} {'Net Pay':>12} {'ER Liab':>10}")
print(f"  {'-'*120}")

grand_gross = 0
grand_taxes = 0
grand_net = 0
grand_er = 0

for e in employees:
    print(f"  {e['name']:<35} {e['dept']:<30} {e['hours']:>7.0f} ${e['gross']:>10,.2f} ${e['ee_taxes']:>8,.2f} ${e['net']:>10,.2f} ${e['er_liability']:>8,.2f}")
    grand_gross += e['gross']
    grand_taxes += e['ee_taxes']
    grand_net += e['net']
    grand_er += e['er_liability']

print(f"  {'-'*120}")
print(f"  {'TOTALS':<35} {'':30} {'':>7} ${grand_gross:>10,.2f} ${grand_taxes:>8,.2f} ${grand_net:>10,.2f} ${grand_er:>8,.2f}")

print(f"\n\n  KEY VERIFICATION:")
print(f"  Total Gross Earnings:     ${grand_gross:>12,.2f}")
print(f"  Total Employee Taxes:     ${grand_taxes:>12,.2f}")
print(f"  Total Net Pay:            ${grand_net:>12,.2f}")
print(f"  Total Employer Liability: ${grand_er:>12,.2f}")
print(f"  Gross - EE Tax - Net:     ${grand_gross - grand_taxes - grand_net:>12,.2f}  (should be ~deductions)")
print(f"  Gross + ER Liability:     ${grand_gross + grand_er:>12,.2f}  (total cost to company)")

print(f"\n\n  MODEL COMPARISON:")
print(f"  Model W-2 total:          $  1,103,125")
print(f"  ADP Gross Earnings:       ${grand_gross:>12,.2f}")
print(f"  Difference:               ${grand_gross - 1_103_125:>12,.2f}")

print(f"\n  Model Payroll Tax:        $    310,876")
print(f"  ADP Employer Liability:   ${grand_er:>12,.2f}")
print(f"  Difference:               ${grand_er - 310_876:>12,.2f}")

# Check payment dates for any 2026 entries
print(f"\n\n  PAYMENT DATE AUDIT:")
all_pay_dates = []
dates_2026 = []
for e in employees:
    for d in e['payment_dates']:
        try:
            for fmt in ['%m/%d/%Y', '%Y-%m-%d']:
                try:
                    dt = datetime.strptime(d, fmt)
                    all_pay_dates.append(dt)
                    if dt.year >= 2026:
                        dates_2026.append((e['name'], d))
                    break
                except:
                    pass
        except:
            pass

if all_pay_dates:
    print(f"  Earliest payment: {min(all_pay_dates).strftime('%Y-%m-%d')}")
    print(f"  Latest payment:   {max(all_pay_dates).strftime('%Y-%m-%d')}")
    by_year = {}
    for d in all_pay_dates:
        by_year[d.year] = by_year.get(d.year, 0) + 1
    print(f"  Payments by year: {by_year}")

if dates_2026:
    print(f"\n  *** WARNING: 2026 PAYMENTS FOUND ***")
    for name, d in dates_2026:
        print(f"    {name}: {d}")
else:
    print(f"  ✓ No 2026 payment dates found")

# Per-person comparison with model
print(f"\n\n  PER-PERSON: ADP GROSS vs MODEL AMOUNT")
print(f"  {'ADP Name':<35} {'ADP Gross':>12} {'Model Amt':>12} {'Diff':>10}")
print(f"  {'-'*75}")

model_amounts = {
    "Acker, Emily": 10_544,
    "Alvarez, Joel": 27_076,
    "Anzalone, Eric": 84_949,
    "Brock, Aimee": 24_423,
    "Brooke, Debby": 7_928,
    "Calkins, Alexandria": 18_436,
    "Calkins, Stephen": 164_960,
    "Carcamo, Andrea": 7_614,
    "Carey, Kyle": 55_978,
    "Copeland, Breanna": 11_719,
    "Davin, Jamie": 10_465,
    "Freeman, Paul": 0,  # Partner, not W-2
    "Freeman, Tracy": 12_889,
    "Froio, James": 127_215,
    "Goodwin, Nicole": 53_342,  # 39,933 + 13,409
    "Harris, Kaitlin": 84_910,
    "Holmes, Ashley": 725,
    "Kohler, Heather": 14_851,
    "McDonell, Hagen": 56_784,
    "Morte, Julia": 20_934,
    "Rankin, Dana": 9_952,
    "Rivera, Anneliese": 4_132,
    "Rocha Moreira, Marcello": 33_280,
    "Simpson, Juliet": 9_263,
    "Walsh, Laura": 124_065,
    "Wiggins, Chad": 108_257,
    "Wingard, Mariette": 0,  # Not in model?
    "Wolf, Emily": 18_434,
}

total_model_match = 0
total_adp_match = 0
for e in employees:
    # Try to match by last name
    adp_name = e['name']
    model_amt = None
    for mname, mamt in model_amounts.items():
        if mname.split(',')[0].lower() in adp_name.lower():
            model_amt = mamt
            break
    
    if model_amt is not None:
        diff = e['gross'] - model_amt
        flag = " ***" if abs(diff) > 100 else ""
        print(f"  {adp_name:<35} ${e['gross']:>10,.0f} ${model_amt:>10,} ${diff:>8,.0f}{flag}")
        total_model_match += model_amt
        total_adp_match += e['gross']
    else:
        print(f"  {adp_name:<35} ${e['gross']:>10,.0f} {'NOT IN MODEL':>12}")
        total_adp_match += e['gross']

print(f"  {'-'*75}")
print(f"  {'TOTALS':<35} ${total_adp_match:>10,.0f} ${total_model_match:>10,} ${total_adp_match - total_model_match:>8,.0f}")

wb.close()
