"""
PulseOne Software COGS & Expense — Three-Bucket Classification
==============================================================
Classifies every line item in 2025 PulseOne Software COGS and Expense.xlsx into:
  1. Client-Delivered Software COGS  — resale to a specific client
  2. Help Desk / Service Delivery COGS — tools used by technicians to deliver services
  3. Internal SG&A — tools used by PulseOne staff for internal operations

Classification rules are based on:
  - Owner-confirmed classifications (architectural-invariants.md)
  - Service Leadership BIC chart-of-accounts logic
  - Vendor function analysis (what the tool actually does)

Outputs:
  - data/financial/Software_COGS_Classified.xlsx  (annotated workbook)
  - docs/analysis/software_classification_summary.json  (summary stats)
"""

import openpyxl
import json
import os
from collections import defaultdict

XLSX_PATH = "/home/ubuntu/bic-work/data/financial/2025 PulseOne Software COGS and Expense.xlsx"
OUT_XLSX  = "/home/ubuntu/bic-work/data/financial/Software_COGS_Classified.xlsx"
OUT_JSON  = "/home/ubuntu/bic-work/docs/analysis/software_classification_summary.json"

# ─────────────────────────────────────────────────────────────────────────────
# CLASSIFICATION RULES
# Priority order: CONFIRMED > FUNCTION-BASED > DEFAULT-BY-SOURCE-SHEET
# ─────────────────────────────────────────────────────────────────────────────

# Bucket labels
B1 = "Client-Delivered Software COGS"        # resale / client-specific
B2 = "Help Desk / Service Delivery COGS"     # technician delivery tools
B3 = "Internal SG&A"                         # internal ops / admin

# Confidence levels
CONFIRMED = "Confirmed (owner)"
ANALYSIS  = "Analysis"
INFERRED  = "Inferred"

# ── Vendor-level override map ─────────────────────────────────────────────────
# Key: vendor name (case-insensitive, partial match allowed)
# Value: (bucket, confidence, rationale)
VENDOR_RULES = {
    # ── CONFIRMED BY OWNER (April 20, 2026) ──────────────────────────────────
    "flexis":           (B2, CONFIRMED, "NOC/Help Desk augmentation — owner confirmed COGS"),
    "amazon web services": (B2, CONFIRMED, "Client-specific AWS environments — owner confirmed COGS"),
    "aws":              (B2, CONFIRMED, "Client-specific AWS environments — owner confirmed COGS"),
    "samurai":          (B2, CONFIRMED, "Service delivery tool — owner confirmed COGS"),

    # ── CONFIRMED BY PRIOR ANALYSIS (April 17–20, 2026) ─────────────────────
    "wasabi":           (B2, ANALYSIS, "Client backup delivery — reclassify from SGA to COGS"),
    "duo security":     (B2, ANALYSIS, "Client MFA delivery — reclassify from SGA to COGS"),
    "qualys":           (B2, ANALYSIS, "Client vulnerability scanning — reclassify from SGA to COGS"),
    "barracuda":        (B2, ANALYSIS, "Managed security/email filtering delivery platform"),
    "skout":            (B2, ANALYSIS, "Managed security delivery platform (Barracuda/Skout)"),
    "connectwise":      (B2, ANALYSIS, "PSA/RMM — primary service delivery platform"),

    # ── CLIENT-DELIVERED SOFTWARE COGS (Resale) ──────────────────────────────
    "synnex":           (B1, ANALYSIS, "Distribution resale — Synnex Corp product purchases"),
    "td synnex":        (B1, ANALYSIS, "Distribution resale — TD Synnex product purchases"),
    "ingram micro":     (B1, ANALYSIS, "Distribution resale — Ingram Micro product purchases"),
    "ingram":           (B1, ANALYSIS, "Distribution resale — Ingram Micro product purchases"),
    "microsoft":        (B1, ANALYSIS, "Microsoft licensing resale (M365, Azure, etc.)"),
    "microsoft corporation": (B1, ANALYSIS, "Microsoft licensing resale"),
    "datto":            (B1, ANALYSIS, "Backup/DR product resale to clients"),
    "veeam":            (B1, ANALYSIS, "Backup product resale to clients"),
    "vmware":           (B1, ANALYSIS, "Virtualization product resale to clients"),
    "cisco":            (B1, ANALYSIS, "Network hardware/software resale"),
    "fortinet":         (B1, ANALYSIS, "Network security product resale"),
    "sonicwall":        (B1, ANALYSIS, "Network security product resale"),
    "watchguard":       (B1, ANALYSIS, "Network security product resale"),
    "sophos":           (B1, ANALYSIS, "Endpoint security product resale"),
    "sentinelone":      (B1, ANALYSIS, "Endpoint security product resale"),
    "crowdstrike":      (B1, ANALYSIS, "Endpoint security product resale"),
    "webroot":          (B1, ANALYSIS, "Endpoint security product resale"),
    "malwarebytes":     (B1, ANALYSIS, "Endpoint security product resale"),
    "acronis":          (B1, ANALYSIS, "Backup product resale"),
    "carbonite":        (B1, ANALYSIS, "Backup product resale"),
    "n-able":           (B1, ANALYSIS, "RMM/backup product resale"),
    "kaseya":           (B1, ANALYSIS, "RMM product resale"),
    "auvik":            (B1, ANALYSIS, "Network monitoring product resale"),
    "liongard":         (B1, ANALYSIS, "Configuration management product resale"),
    "huntress":         (B1, ANALYSIS, "MDR product resale to clients"),
    "cove":             (B1, ANALYSIS, "Backup product resale (N-able Cove)"),
    "dropsuite":        (B1, ANALYSIS, "Email backup product resale"),
    "proofpoint":       (B1, ANALYSIS, "Email security product resale"),
    "mimecast":         (B1, ANALYSIS, "Email security product resale"),
    "knowbe4":          (B1, ANALYSIS, "Security awareness training resale"),
    "proofpoint":       (B1, ANALYSIS, "Email security product resale"),
    "nerdio":           (B1, ANALYSIS, "Azure Virtual Desktop management resale"),
    "pax8":             (B1, ANALYSIS, "Cloud distribution resale"),
    "sherweb":          (B1, ANALYSIS, "Cloud distribution resale"),
    "also cloud":       (B1, ANALYSIS, "Cloud distribution resale"),
    "tech data":        (B1, ANALYSIS, "Distribution resale"),
    "d&h":              (B1, ANALYSIS, "Distribution resale"),
    "avepoint":         (B1, ANALYSIS, "Microsoft 365 backup/governance resale"),
    "spanning":         (B1, ANALYSIS, "SaaS backup product resale"),
    "backupify":        (B1, ANALYSIS, "SaaS backup product resale"),
    "zix":              (B1, ANALYSIS, "Email encryption product resale"),
    "abnormal":         (B1, ANALYSIS, "Email security product resale"),
    "ironscales":       (B1, ANALYSIS, "Email security product resale"),
    "cofense":          (B1, ANALYSIS, "Phishing simulation product resale"),

    # ── HELP DESK / SERVICE DELIVERY COGS ────────────────────────────────────
    "pulseway":         (B2, ANALYSIS, "RMM platform — service delivery"),
    "ninja":            (B2, ANALYSIS, "RMM platform — service delivery"),
    "ninjarmm":         (B2, ANALYSIS, "RMM platform — service delivery"),
    "atera":            (B2, ANALYSIS, "RMM/PSA platform — service delivery"),
    "syncro":           (B2, ANALYSIS, "RMM/PSA platform — service delivery"),
    "halo":             (B2, ANALYSIS, "PSA platform — service delivery"),
    "halopsa":          (B2, ANALYSIS, "PSA platform — service delivery"),
    "autotask":         (B2, ANALYSIS, "PSA platform — service delivery"),
    "servicenow":       (B2, ANALYSIS, "ITSM platform — service delivery"),
    "jira":             (B2, ANALYSIS, "Ticketing — service delivery"),
    "atlassian":        (B2, INFERRED, "Jira/Confluence — likely service delivery ticketing; confirm internal vs client"),
    "teamviewer":       (B2, ANALYSIS, "Remote access — service delivery"),
    "splashtop":        (B2, ANALYSIS, "Remote access — service delivery"),
    "screenconnect":    (B2, ANALYSIS, "Remote access — service delivery"),
    "bomgar":           (B2, ANALYSIS, "Remote access — service delivery"),
    "beyondtrust":      (B2, ANALYSIS, "Remote access — service delivery"),
    "logmein":          (B2, ANALYSIS, "Remote access — service delivery"),
    "anydesk":          (B2, ANALYSIS, "Remote access — service delivery"),
    "n-central":        (B2, ANALYSIS, "RMM — service delivery"),
    "it glue":          (B2, ANALYSIS, "Documentation platform — service delivery"),
    "itglue":           (B2, ANALYSIS, "Documentation platform — service delivery"),
    "hudu":             (B2, ANALYSIS, "Documentation platform — service delivery"),
    "passportal":       (B2, ANALYSIS, "Password/documentation — service delivery"),
    "myglue":           (B2, ANALYSIS, "Client documentation portal — service delivery"),
    "dark web":         (B2, ANALYSIS, "Dark web monitoring — service delivery"),
    "id agent":         (B2, ANALYSIS, "Dark web monitoring — service delivery"),
    "lifelock":         (B2, ANALYSIS, "Identity monitoring — service delivery"),
    "threatlocker":     (B2, ANALYSIS, "Zero-trust security — service delivery"),
    "blackpoint":       (B2, ANALYSIS, "MDR — service delivery"),
    "guardz":           (B2, ANALYSIS, "SMB security platform — service delivery"),
    "perch":            (B2, ANALYSIS, "SIEM/SOC — service delivery"),
    "blumira":          (B2, ANALYSIS, "SIEM — service delivery"),
    "coreview":         (B2, ANALYSIS, "M365 management — service delivery"),
    "simplesat":        (B2, ANALYSIS, "Client satisfaction surveys — service delivery"),
    "nps":              (B2, ANALYSIS, "Client satisfaction — service delivery"),
    "hpbuy":            (B2, ANALYSIS, "HP procurement for client delivery"),
    "dell":             (B2, ANALYSIS, "Hardware procurement for client delivery"),
    "lenovo":           (B2, ANALYSIS, "Hardware procurement for client delivery"),

    # ── INTERNAL SG&A ─────────────────────────────────────────────────────────
    "openai":           (B3, ANALYSIS, "Internal AI tool — SG&A"),
    "anthropic":        (B3, ANALYSIS, "Internal AI tool — SG&A"),
    "claude":           (B3, ANALYSIS, "Internal AI tool — SG&A"),
    "openrouter":       (B3, ANALYSIS, "Internal AI routing — SG&A"),
    "cursor":           (B3, ANALYSIS, "Internal dev IDE — SG&A"),
    "replit":           (B3, ANALYSIS, "Internal dev environment — SG&A"),
    "railway":          (B3, ANALYSIS, "Internal hosting/dev — SG&A"),
    "github":           (B3, ANALYSIS, "Internal code repository — SG&A"),
    "scrapfly":         (B3, ANALYSIS, "Internal web scraping — SG&A"),
    "iron software":    (B3, ANALYSIS, "Internal dev library — SG&A"),
    "calendly":         (B3, ANALYSIS, "Internal scheduling — SG&A (redundant with TimeZest)"),
    "timezest":         (B3, ANALYSIS, "Scheduling integration with ConnectWise — borderline COGS/SGA; keep as SGA"),
    "repscheduler":     (B3, ANALYSIS, "Sales scheduling tool — SG&A"),
    "monday":           (B3, ANALYSIS, "Internal project management — SG&A"),
    "monday.com":       (B3, ANALYSIS, "Internal project management — SG&A"),
    "figma":            (B3, ANALYSIS, "Internal design tool — SG&A"),
    "adobe":            (B3, ANALYSIS, "Internal creative suite — SG&A"),
    "canva":            (B3, ANALYSIS, "Internal design tool — SG&A"),
    "patreon":          (B3, ANALYSIS, "Non-business subscription — SG&A (review for cancellation)"),
    "quicken":          (B3, ANALYSIS, "Personal finance tool — SG&A (review for cancellation)"),
    "vevox":            (B3, ANALYSIS, "Meeting polling tool — SG&A (redundant)"),
    "learndash":        (B3, ANALYSIS, "LMS — SG&A (internal training)"),
    "avalara":          (B3, ANALYSIS, "Tax compliance — SG&A"),
    "eversign":         (B3, ANALYSIS, "E-signature — SG&A"),
    "docusign":         (B3, ANALYSIS, "E-signature — SG&A"),
    "godaddy":          (B3, INFERRED, "Domain/hosting — likely internal; confirm if client-hosted"),
    "google":           (B3, ANALYSIS, "Google Workspace — internal productivity"),
    "google llc":       (B3, ANALYSIS, "Google Workspace — internal productivity"),
    "zoom":             (B3, ANALYSIS, "Internal communications — SG&A"),
    "slack":            (B3, ANALYSIS, "Internal communications — SG&A"),
    "teams":            (B3, ANALYSIS, "Internal communications — SG&A"),
    "dropbox":          (B3, ANALYSIS, "Internal file storage — SG&A"),
    "box":              (B3, ANALYSIS, "Internal file storage — SG&A"),
    "1password":        (B3, ANALYSIS, "Internal password manager — SG&A"),
    "lastpass":         (B3, ANALYSIS, "Internal password manager — SG&A"),
    "bitwarden":        (B3, ANALYSIS, "Internal password manager — SG&A"),
    "zapier":           (B3, ANALYSIS, "Internal automation — SG&A"),
    "make":             (B3, ANALYSIS, "Internal automation — SG&A"),
    "hubspot":          (B3, ANALYSIS, "CRM — SG&A"),
    "salesforce":       (B3, ANALYSIS, "CRM — SG&A"),
    "constant contact": (B3, ANALYSIS, "Email marketing — SG&A"),
    "mailchimp":        (B3, ANALYSIS, "Email marketing — SG&A"),
    "linkedin":         (B3, ANALYSIS, "Marketing/recruiting — SG&A"),
    "indeed":           (B3, ANALYSIS, "Recruiting — SG&A"),
    "bamboohr":         (B3, ANALYSIS, "HR platform — SG&A"),
    "rippling":         (B3, ANALYSIS, "HR/payroll platform — SG&A"),
    "gusto":            (B3, ANALYSIS, "Payroll — SG&A"),
    "adp":              (B3, ANALYSIS, "Payroll — SG&A"),
    "quickbooks":       (B3, ANALYSIS, "Accounting — SG&A"),
    "intuit":           (B3, ANALYSIS, "Accounting/QuickBooks — SG&A"),
    "xero":             (B3, ANALYSIS, "Accounting — SG&A"),
    "bill.com":         (B3, ANALYSIS, "AP automation — SG&A"),
    "expensify":        (B3, ANALYSIS, "Expense management — SG&A"),
    "concur":           (B3, ANALYSIS, "Expense management — SG&A"),
    "azure":            (B3, INFERRED, "Microsoft Azure — confirm if client-specific (COGS) or internal infra (SGA)"),
    "microsoft azure":  (B3, INFERRED, "Microsoft Azure — confirm if client-specific (COGS) or internal infra (SGA)"),
}

# ── Source-sheet default rules ────────────────────────────────────────────────
# If no vendor match, fall back to the sheet/category the row came from
SHEET_DEFAULTS = {
    "Agreement License Cost": (B2, INFERRED, "Default: Agreement License Cost sheet — service delivery tool"),
    "Purchases - Software for Resale": (B1, INFERRED, "Default: Software Resale sheet — client-delivered product"),
    "Computer & Internet Expenses": (B3, INFERRED, "Default: Computer & Internet SGA sheet — internal tool"),
}

def classify_vendor(vendor_name: str, sheet_name: str) -> tuple:
    """Return (bucket, confidence, rationale) for a given vendor and sheet."""
    if not vendor_name:
        bucket, conf, rat = SHEET_DEFAULTS.get(sheet_name, (B3, INFERRED, "Unknown sheet — defaulted to SGA"))
        return bucket, conf, rat

    vl = vendor_name.lower().strip()

    # Check vendor rules (longest match wins to handle "microsoft" vs "microsoft azure")
    best_match_len = 0
    best_rule = None
    for key, rule in VENDOR_RULES.items():
        if key in vl and len(key) > best_match_len:
            best_match_len = len(key)
            best_rule = rule

    if best_rule:
        return best_rule

    # Fall back to sheet default
    bucket, conf, rat = SHEET_DEFAULTS.get(sheet_name, (B3, INFERRED, "No vendor match — defaulted to SGA"))
    return bucket, conf, rat


def run():
    wb_in = openpyxl.load_workbook(XLSX_PATH)
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)  # remove default sheet

    summary = {
        "total_rows": 0,
        "total_amount": 0.0,
        "by_bucket": {B1: {"count": 0, "amount": 0.0, "vendors": {}},
                      B2: {"count": 0, "amount": 0.0, "vendors": {}},
                      B3: {"count": 0, "amount": 0.0, "vendors": {}}},
        "by_sheet": {},
        "unmatched_vendors": {},
        "needs_confirmation": [],
        "sheet_details": {}
    }

    # ── Process each sheet ────────────────────────────────────────────────────
    for sheet_name in wb_in.sheetnames:
        ws_in = wb_in[sheet_name]
        ws_out = wb_out.create_sheet(title=sheet_name[:31])

        # Read header row
        headers = [cell.value for cell in ws_in[1]]

        # Add classification columns
        new_headers = headers + ["Bucket", "Confidence", "Rationale"]
        ws_out.append(new_headers)

        # Style header
        from openpyxl.styles import PatternFill, Font, Alignment
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws_out[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Bucket fill colors
        bucket_fills = {
            B1: PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),  # green
            B2: PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),  # blue
            B3: PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),  # yellow
        }
        confirm_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")  # orange for needs confirmation

        sheet_summary = {"count": 0, "amount": 0.0,
                         B1: 0.0, B2: 0.0, B3: 0.0}

        # Find vendor and amount column indices
        vendor_col = None
        amount_col = None
        for i, h in enumerate(headers):
            if h and "name" in str(h).lower():
                vendor_col = i
            if h and ("amount" in str(h).lower() or "total" in str(h).lower() or "cost" in str(h).lower()):
                if amount_col is None:
                    amount_col = i

        # Fallback: try column B for vendor, column D or E for amount
        if vendor_col is None:
            vendor_col = 1  # column B (0-indexed)
        if amount_col is None:
            amount_col = 3  # column D (0-indexed)

        for row_idx, row in enumerate(ws_in.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue

            vendor = str(row[vendor_col]).strip() if row[vendor_col] else ""
            if vendor in ("None", "nan", ""):
                vendor = ""

            # Try to get amount
            amount = 0.0
            for ci in range(len(row)-1, -1, -1):
                val = row[ci]
                if isinstance(val, (int, float)) and val != 0:
                    amount = float(val)
                    break

            bucket, confidence, rationale = classify_vendor(vendor, sheet_name)

            new_row = list(row) + [bucket, confidence, rationale]
            ws_out.append(new_row)

            # Apply row fill
            fill = bucket_fills.get(bucket, PatternFill())
            if confidence == INFERRED:
                fill = confirm_fill
            data_row = ws_out[ws_out.max_row]
            for cell in data_row:
                cell.fill = fill

            # Update summary
            summary["total_rows"] += 1
            summary["total_amount"] += amount
            summary["by_bucket"][bucket]["count"] += 1
            summary["by_bucket"][bucket]["amount"] += amount

            # Vendor rollup
            if vendor:
                vkey = vendor[:50]
                if vkey not in summary["by_bucket"][bucket]["vendors"]:
                    summary["by_bucket"][bucket]["vendors"][vkey] = 0.0
                summary["by_bucket"][bucket]["vendors"][vkey] += amount

            sheet_summary["count"] += 1
            sheet_summary["amount"] += amount
            sheet_summary[bucket] = sheet_summary.get(bucket, 0.0) + amount

            if confidence == INFERRED:
                summary["needs_confirmation"].append({
                    "vendor": vendor,
                    "sheet": sheet_name,
                    "amount": amount,
                    "current_bucket": bucket,
                    "rationale": rationale
                })

        summary["by_sheet"][sheet_name] = sheet_summary
        summary["sheet_details"][sheet_name] = {
            "vendor_col": vendor_col,
            "amount_col": amount_col,
            "headers": headers
        }

    # ── Add summary sheet ─────────────────────────────────────────────────────
    ws_sum = wb_out.create_sheet(title="CLASSIFICATION SUMMARY", index=0)
    ws_sum.append(["PulseOne Software COGS & Expense — Three-Bucket Classification"])
    ws_sum.append(["Source: 2025 PulseOne Software COGS and Expense.xlsx"])
    ws_sum.append(["Classified: April 20, 2026"])
    ws_sum.append([])
    ws_sum.append(["BUCKET", "LINE COUNT", "TOTAL AMOUNT", "% OF TOTAL"])

    total = summary["total_amount"] if summary["total_amount"] else 1
    for b in [B1, B2, B3]:
        amt = summary["by_bucket"][b]["amount"]
        cnt = summary["by_bucket"][b]["count"]
        ws_sum.append([b, cnt, round(amt, 2), f"{amt/total*100:.1f}%"])

    ws_sum.append([])
    ws_sum.append(["TOTAL", summary["total_rows"], round(summary["total_amount"], 2), "100.0%"])
    ws_sum.append([])
    ws_sum.append(["ITEMS NEEDING OWNER CONFIRMATION (INFERRED classification)", len(summary["needs_confirmation"])])
    ws_sum.append([])
    ws_sum.append(["TOP VENDORS BY BUCKET"])
    ws_sum.append([])

    for b in [B1, B2, B3]:
        ws_sum.append([f"--- {b} ---"])
        vendors = summary["by_bucket"][b]["vendors"]
        sorted_v = sorted(vendors.items(), key=lambda x: x[1], reverse=True)[:15]
        for vname, vamt in sorted_v:
            ws_sum.append(["", vname, round(vamt, 2)])
        ws_sum.append([])

    # Save workbook
    wb_out.save(OUT_XLSX)
    print(f"Saved classified workbook: {OUT_XLSX}")

    # Save JSON summary
    # Convert nested dicts for JSON serialization
    json_summary = {
        "total_rows": summary["total_rows"],
        "total_amount": round(summary["total_amount"], 2),
        "by_bucket": {
            b: {
                "count": summary["by_bucket"][b]["count"],
                "amount": round(summary["by_bucket"][b]["amount"], 2),
                "top_vendors": sorted(
                    [(k, round(v, 2)) for k, v in summary["by_bucket"][b]["vendors"].items()],
                    key=lambda x: x[1], reverse=True
                )[:20]
            }
            for b in [B1, B2, B3]
        },
        "by_sheet": {
            k: {kk: (round(vv, 2) if isinstance(vv, float) else vv)
                for kk, vv in v.items() if kk != "headers"}
            for k, v in summary["by_sheet"].items()
        },
        "needs_confirmation_count": len(summary["needs_confirmation"]),
        "needs_confirmation": summary["needs_confirmation"][:50]
    }

    os.makedirs(os.path.dirname(OUT_JSON), exist_ok=True)
    with open(OUT_JSON, "w") as f:
        json.dump(json_summary, f, indent=2)
    print(f"Saved JSON summary: {OUT_JSON}")

    # Print console summary
    print("\n" + "="*70)
    print("CLASSIFICATION SUMMARY")
    print("="*70)
    for b in [B1, B2, B3]:
        amt = summary["by_bucket"][b]["amount"]
        cnt = summary["by_bucket"][b]["count"]
        print(f"\n{b}")
        print(f"  Lines: {cnt}  |  Amount: ${amt:,.2f}  |  {amt/total*100:.1f}%")
        vendors = summary["by_bucket"][b]["vendors"]
        sorted_v = sorted(vendors.items(), key=lambda x: x[1], reverse=True)[:8]
        for vname, vamt in sorted_v:
            print(f"    {vname:<45} ${vamt:>12,.2f}")

    print(f"\nTotal rows processed: {summary['total_rows']}")
    print(f"Total amount: ${summary['total_amount']:,.2f}")
    print(f"Items needing confirmation: {len(summary['needs_confirmation'])}")

    return json_summary


if __name__ == "__main__":
    result = run()
