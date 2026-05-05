#!/usr/bin/env python3
"""
Extract the G&A operating expense detail from the QB P&L file.
"""

import openpyxl

wb = openpyxl.load_workbook(
    '/home/ubuntu/projects/pulseone-best-in-class-audit-b4407871/2025 Profit and Loss POTS_analysis.xlsx',
    data_only=True
)

ws = wb.active
print(f"Dimensions: {ws.max_row} rows x {ws.max_column} cols")

# Print ALL rows with any data
print("\nALL ROWS WITH DATA:")
print("=" * 120)
for row_idx in range(1, ws.max_row + 1):
    row_data = []
    for col_idx in range(1, ws.max_column + 1):
        v = ws.cell(row=row_idx, column=col_idx).value
        if v is not None:
            row_data.append((col_idx, v))
    if row_data:
        # Show first item and check if it's a label
        first_val = row_data[0][1]
        if isinstance(first_val, str):
            # This is a label row - show label and numeric values
            nums = [(c, v) for c, v in row_data if isinstance(v, (int, float))]
            if nums:
                # Find what looks like an annual total (largest absolute value or sum of 12)
                num_vals = [v for _, v in nums]
                if len(num_vals) >= 12:
                    annual = sum(num_vals[:12])
                    print(f"  {first_val:<40} Annual(sum12): ${annual:>12,.2f}  (vals: {len(num_vals)})")
                else:
                    print(f"  {first_val:<40} Vals: {num_vals}")
            else:
                print(f"  {first_val:<40} (no numbers)")
        else:
            # Numeric first column
            print(f"  Row {row_idx}: starts with {first_val}")
